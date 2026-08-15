import streamlit as st
import yfinance as yf
import pandas as pd
import numpy as np
import time
import requests
import io
from plotly.subplots import make_subplots
import plotly.graph_objects as go

st.set_page_config(page_title="Live Stock Dashboard", layout="wide")

st.title("📈 Live Stock Dashboard")
st.caption("Latest Price | 3-Period EMA | 21-Period WMA | 9-Period RSI")

# Major market-cap based & sectoral indexes (verified Yahoo Finance symbols)
MAJOR_INDEXES = {
    "NIFTY 50": "^NSEI",
    "NIFTY NEXT 50": "^NSMIDCP",
    "NIFTY 100": "^CNX100",
    "NIFTY 200": "^CNX200",
    "NIFTY 500": "^CRSLDX",
    "NIFTY MIDCAP 50": "^NSEMDCP50",
    "NIFTY MIDCAP 100": "NIFTY_MIDCAP_100.NS",
    "NIFTY MIDCAP 150": "NIFTYMIDCAP150.NS",
    "NIFTY SMLCAP 50": "NIFTYSMLCAP50.NS",
    "NIFTY SMLCAP 100": "^CNXSC",
    "NIFTY SMLCAP 250": "NIFTYSMLCAP250.NS",
    "S&P BSE SENSEX": "^BSESN",
    "NIFTY BANK": "^NSEBANK",
    "NIFTY IT": "^CNXIT",
    "NIFTY AUTO": "^CNXAUTO",
    "NIFTY METAL": "^CNXMETAL",
    "NIFTY FMCG": "^CNXFMCG",
    "NIFTY PHARMA": "^CNXPHARMA",
    "NIFTY ENERGY": "^CNXENERGY",
    "NIFTY REALTY": "^CNXREALTY",
    "NIFTY MEDIA": "^CNXMEDIA",
    "NIFTY PSU BANK": "^CNXPSUBANK",
    "NIFTY FIN SERVICE": "NIFTY_FIN_SERVICE.NS",
    "NIFTY INFRA": "^CNXINFRA",
    "NIFTY COMMODITIES": "^CNXCMDT",
}

# Small backup list used only if the live NSE fetch below fails/is blocked
FALLBACK_SYMBOLS = [
    "RELIANCE", "TCS", "HDFCBANK", "ICICIBANK", "INFY", "HINDUNILVR", "ITC",
    "SBIN", "BHARTIARTL", "KOTAKBANK", "LT", "AXISBANK", "BAJFINANCE",
    "ASIANPAINT", "MARUTI", "HCLTECH", "SUNPHARMA", "TITAN", "ULTRACEMCO",
    "NESTLEIND"
]


@st.cache_data(ttl=86400, show_spinner=False)
def get_nifty500_details():
    """Fetch the live Nifty 500 constituent list + company names from NSE
    (the CSV already includes names, so no per-stock Yahoo lookup is
    needed -- 500 individual lookups would be far too slow).
    Falls back to a small known list if NSE blocks the request. This is
    just the constituent list (fast) -- separate from the full price scan."""
    url = "https://nsearchives.nseindia.com/content/indices/ind_nifty500list.csv"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0 Safari/537.36",
        "Accept-Language": "en-US,en;q=0.9",
    }
    try:
        session = requests.Session()
        session.headers.update(headers)
        session.get("https://www.nseindia.com", timeout=10)  # sets cookies
        resp = session.get(url, timeout=15)
        resp.raise_for_status()
        df = pd.read_csv(io.StringIO(resp.text))
        symbols = df["Symbol"].astype(str).str.strip().tolist()
        names = dict(zip(
            df["Symbol"].astype(str).str.strip(),
            df["Company Name"].astype(str).str.strip()
        ))
        return symbols, names
    except Exception:
        return FALLBACK_SYMBOLS, {}


# ---------- Sidebar inputs ----------
st.sidebar.header("Settings")
symbol = st.sidebar.text_input(
    "Enter Stock Symbol (Yahoo Finance format)",
    value=st.session_state.get("symbol_input", "^NSEI"),
    help="NSE stocks need '.NS' suffix. Example: TCS.NS, INFY.NS, RELIANCE.NS",
    key="symbol_input"
)

st.sidebar.subheader("📋 Copy a ticker")
with st.sidebar.expander("Indexes"):
    _idx_search = st.text_input("🔍 Search index", key="idx_search", placeholder="e.g. bank, next 50")
    _idx_options = list(MAJOR_INDEXES.items())  # (name, ticker) pairs
    if _idx_search:
        _idx_options = [
            (n, t) for n, t in _idx_options
            if _idx_search.lower() in n.lower() or _idx_search.lower() in t.lower()
        ]
    if _idx_options:
        _idx_pick = st.selectbox(
            "Select", [f"{n}: {t}" for n, t in _idx_options], key="idx_pick",
            label_visibility="collapsed"
        )
        _idx_pick_ticker = _idx_pick.rsplit(": ", 1)[-1]
        st.code(_idx_pick_ticker, language=None)  # only the ticker -- nothing else to copy
    else:
        st.caption("No match found.")

with st.sidebar.expander("Nifty 500 Stocks"):
    _stock_search = st.text_input("🔍 Search stock", key="stock_search", placeholder="e.g. reliance, TCS")
    with st.spinner("Loading Nifty 500 list from NSE..."):
        _n500_symbols, _n500_names = get_nifty500_details()
    _stock_options = [(_n500_names.get(s, s), f"{s}.NS") for s in _n500_symbols]
    if _stock_search:
        _stock_options = [
            (n, t) for n, t in _stock_options
            if _stock_search.lower() in n.lower() or _stock_search.lower() in t.lower()
        ]
    if _stock_options:
        _stock_pick = st.selectbox(
            "Select", [f"{n}: {t}" for n, t in _stock_options], key="stock_pick",
            label_visibility="collapsed"
        )
        _stock_pick_ticker = _stock_pick.rsplit(": ", 1)[-1]
        st.code(_stock_pick_ticker, language=None)  # only the ticker -- nothing else to copy
    else:
        st.caption("No match found.")
        st.caption("No match found.")

# Yahoo Finance does not natively provide 10m / 2h / 3h / 4h candles.
# We fetch the closest available candle size and combine ("resample")
# it into the requested duration ourselves.
INTERVAL_CONFIG = {
    "1m":  {"yf_interval": "1m",  "period": "7d",   "resample": None},
    "5m":  {"yf_interval": "5m",  "period": "60d",  "resample": None},
    "10m": {"yf_interval": "5m",  "period": "60d",  "resample": "10min"},
    "15m": {"yf_interval": "15m", "period": "60d",  "resample": None},
    "30m": {"yf_interval": "30m", "period": "60d",  "resample": None},
    "1h":  {"yf_interval": "60m", "period": "730d", "resample": None},
    "2h":  {"yf_interval": "60m", "period": "730d", "resample": "2h"},
    "3h":  {"yf_interval": "60m", "period": "730d", "resample": "3h"},
    "4h":  {"yf_interval": "60m", "period": "730d", "resample": "4h"},
    "1d":  {"yf_interval": "1d",  "period": "10y",  "resample": None},
    "1wk": {"yf_interval": "1wk", "period": "max",  "resample": None},
    "1mo": {"yf_interval": "1mo", "period": "max",  "resample": None},
}

interval = st.sidebar.selectbox(
    "Candle Interval",
    options=list(INTERVAL_CONFIG.keys()),
    index=list(INTERVAL_CONFIG.keys()).index("1d")
)
refresh_seconds = st.sidebar.number_input(
    "Auto-refresh every (seconds)",
    min_value=10, max_value=300, value=30
)
num_bars = st.sidebar.slider(
    "Candles to show on chart",
    min_value=30, max_value=300, value=45,
    help="Fewer candles = cleaner, less zig-zag chart (like TradingView's default zoomed view)"
)
run = st.sidebar.checkbox("Start Live Updates", value=True)


# ---------- Indicator functions ----------
def calculate_indicators(df):
    df = df.copy()

    # 3-period EMA of Close price
    df["EMA_3"] = df["Close"].ewm(span=3, adjust=False).mean()

    # 21-period WMA (Weighted Moving Average) of Close price
    weights = np.arange(1, 22)  # 1 to 21
    df["WMA_21"] = df["Close"].rolling(21).apply(
        lambda prices: np.dot(prices, weights) / weights.sum(), raw=True
    )

    # 9-period RSI using Wilder's smoothing (same method TradingView uses)
    # -> gives a much smoother line than a simple rolling average.
    delta = df["Close"].diff()
    gain = delta.clip(lower=0)
    loss = -delta.clip(upper=0)
    avg_gain = gain.ewm(alpha=1 / 9, min_periods=9, adjust=False).mean()
    avg_loss = loss.ewm(alpha=1 / 9, min_periods=9, adjust=False).mean()
    rs = avg_gain / avg_loss
    df["RSI_9"] = 100 - (100 / (1 + rs))

    # "Hilega Milega" style: EMA(3) and WMA(21) applied to the RSI itself
    # (not to price) so both lines sit on the same 0-100 scale as RSI.
    df["RSI_EMA_3"] = df["RSI_9"].ewm(span=3, adjust=False).mean()
    df["RSI_WMA_21"] = df["RSI_9"].rolling(21).apply(
        lambda vals: np.dot(vals, weights) / weights.sum(), raw=True
    )

    return df


# ---------- Data fetch ----------
@st.cache_data(ttl=3600, show_spinner=False)
def get_company_name(sym):
    """Look up the friendly company/index name for the selected symbol."""
    try:
        info = yf.Ticker(sym).info
        return info.get("longName") or info.get("shortName") or sym
    except Exception:
        return sym


def fetch_data(symbol, interval):
    cfg = INTERVAL_CONFIG[interval]

    data = yf.download(
        tickers=symbol,
        period=cfg["period"],
        interval=cfg["yf_interval"],
        progress=False
    )

    if data.empty:
        return data

    if isinstance(data.columns, pd.MultiIndex):
        data.columns = data.columns.get_level_values(0)

    if cfg["resample"]:
        # Anchor bins to NSE's 9:15 AM market open (default resample bins
        # start at midnight, which chops candles into uneven/partial pieces
        # and made 2h/3h/4h charts look distorted).
        data = data.resample(cfg["resample"], offset="9h15min").agg({
            "Open": "first",
            "High": "max",
            "Low": "min",
            "Close": "last",
            "Volume": "sum"
        }).dropna(how="all")

    return data


# ---------- Chart (TradingView "Hilega Milega" style) ----------
def build_chart(data, num_bars=45, interval="1d"):
    idx = data.index
    n = len(data)
    x = list(range(n))  # sequential positions -- every candle is equally
    # spaced regardless of real elapsed time, so weekends/off-hours/uneven
    # resampled bins can never create gaps, overlaps, or squeezed candles
    # (this is how TradingView spaces candles too -- only the axis LABELS
    # show the real date/time, adapted to the chosen timeframe).

    is_intraday = interval in ["1m", "5m", "10m", "15m", "30m", "1h", "2h", "3h", "4h"]
    fine_intraday = interval in ["1m", "5m", "10m", "15m", "30m"]

    # Full label (used on hover) and a shorter tick label, both matched to
    # the selected timeframe -- like TradingView adapts label granularity
    # to the chart's zoom/interval.
    if interval in ["1wk", "1mo"]:
        labels = [d.strftime("%b %Y") for d in idx]
        tick_labels = labels
    elif interval == "1d":
        labels = [d.strftime("%d-%b-%Y") for d in idx]
        tick_labels = [d.strftime("%d-%b") for d in idx]
    elif is_intraday:
        labels = [d.strftime("%d-%b-%Y %H:%M") for d in idx]
        if fine_intraday:
            # show just the time, except the first candle of a new day
            # (then show the date too, like TradingView's day dividers)
            tick_labels = []
            prev_day = None
            for d in idx:
                if d.date() != prev_day:
                    tick_labels.append(d.strftime("%d-%b %H:%M"))
                    prev_day = d.date()
                else:
                    tick_labels.append(d.strftime("%H:%M"))
        else:
            tick_labels = [d.strftime("%d-%b %H:%M") for d in idx]
    else:
        labels = [d.strftime("%d-%b-%Y") for d in idx]
        tick_labels = labels

    rsi = data["RSI_9"]
    rsi_wma = data["RSI_WMA_21"]
    rsi_ema = data["RSI_EMA_3"]

    # Masked series so the fill collapses to nothing on the "wrong" side of 50
    rsi_upper = rsi.clip(lower=50)
    rsi_lower = rsi.clip(upper=50)
    midline = pd.Series(50, index=idx)

    fig = make_subplots(
        rows=2, cols=1,
        shared_xaxes=True,
        row_heights=[0.55, 0.45],
        vertical_spacing=0.04,
    )

    # ---- Upper section: Price as a candlestick chart ----
    fig.add_trace(
        go.Candlestick(
            x=x, open=data["Open"], high=data["High"], low=data["Low"], close=data["Close"],
            name="Price", text=labels,
            hovertemplate="%{text}<br>O: %{open:.2f}  H: %{high:.2f}<br>L: %{low:.2f}  C: %{close:.2f}<extra></extra>",
            increasing_line_color="#26A69A", decreasing_line_color="#EF5350"
        ),
        row=1, col=1
    )

    # ---- Lower section: RSI zones (orange above 50, light green below 50) ----
    fig.add_trace(go.Scatter(x=x, y=midline, line=dict(width=0), showlegend=False,
                              hoverinfo="skip"), row=2, col=1)
    fig.add_trace(go.Scatter(x=x, y=rsi_upper, line=dict(width=0), fill="tonexty",
                              fillcolor="rgba(144,238,144,0.35)", name="Overbought (>50)",
                              showlegend=True, hoverinfo="skip"), row=2, col=1)
    fig.add_trace(go.Scatter(x=x, y=midline, line=dict(width=0), showlegend=False,
                              hoverinfo="skip"), row=2, col=1)
    fig.add_trace(go.Scatter(x=x, y=rsi_lower, line=dict(width=0), fill="tonexty",
                              fillcolor="rgba(255,200,120,0.35)", name="Oversold (<50)",
                              showlegend=True, hoverinfo="skip"), row=2, col=1)

    # Midline at 50
    fig.add_trace(
        go.Scatter(x=x, y=midline, mode="lines", line=dict(color="gray", width=1, dash="dot"),
                   name="Midline (50)"), row=2, col=1
    )

    # RSI line itself
    fig.add_trace(
        go.Scatter(x=x, y=rsi, mode="lines", line=dict(color="white", width=1.1),
                   name="RSI (9)", text=labels, hovertemplate="%{text}<br>RSI: %{y:.2f}<extra></extra>"),
        row=2, col=1
    )

    # WMA(21) and EMA(3) of the RSI itself -- same 0-100 scale, like the
    # "Hilega Milega" TradingView indicator (close, 9, 3, 21)
    fig.add_trace(
        go.Scatter(x=x, y=rsi_wma, mode="lines", line=dict(color="red", width=1),
                   name="21 WMA (of RSI)", text=labels,
                   hovertemplate="%{text}<br>WMA: %{y:.2f}<extra></extra>"),
        row=2, col=1
    )
    fig.add_trace(
        go.Scatter(x=x, y=rsi_ema, mode="lines", line=dict(color="violet", width=1),
                   name="3 EMA (of RSI)", text=labels,
                   hovertemplate="%{text}<br>EMA: %{y:.2f}<extra></extra>"),
        row=2, col=1
    )

    fig.update_yaxes(title_text="Price", row=1, col=1, fixedrange=True)
    fig.update_yaxes(title_text="RSI", range=[0, 100], row=2, col=1, fixedrange=True)
    fig.update_xaxes(rangeslider_visible=False, row=1, col=1)

    # Custom tick labels along the x-axis (a subset, so they don't overlap)
    max_ticks = 8
    step = max(1, n // max_ticks)
    tickvals = list(range(0, n, step))
    ticktext = [tick_labels[i] for i in tickvals]
    fig.update_xaxes(tickmode="array", tickvals=tickvals, ticktext=ticktext, row=1, col=1)
    fig.update_xaxes(tickmode="array", tickvals=tickvals, ticktext=ticktext, row=2, col=1)

    # Show only the last `num_bars` candles initially, but keep the full
    # fetched history in the figure so scrolling left reveals real older
    # candles instead of hitting empty space.
    if n > num_bars:
        initial_range = [n - num_bars - 0.5, n - 0.5]
    else:
        initial_range = [-0.5, n - 0.5]
    fig.update_xaxes(range=initial_range, row=1, col=1)
    fig.update_xaxes(range=initial_range, row=2, col=1)

    fig.update_layout(
        height=700,
        template="plotly_dark",
        hovermode="x unified",
        margin=dict(l=10, r=10, t=30, b=10),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        dragmode="pan",  # single-finger drag = scroll left/right, not a distorting zoom-box
    )

    return fig


# ---------- Fetch + display ----------
def fetch_and_display():
    try:
        data = fetch_data(symbol, interval)

        if data.empty:
            st.error("No data found. Check the symbol (e.g. use RELIANCE.NS for NSE stocks).")
            return

        data = calculate_indicators(data)
        latest = data.iloc[-1]

        latest_close = float(latest['Close'])
        col_name = get_company_name(symbol)
        st.subheader(f"📊 {col_name} ({symbol})")
        latest_ema = float(latest['EMA_3'])
        latest_wma = float(latest['WMA_21']) if not pd.isna(latest['WMA_21']) else float('nan')
        latest_rsi = float(latest['RSI_9']) if not pd.isna(latest['RSI_9']) else float('nan')

        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Last Traded Price", f"{latest_close:.2f}")
        col2.metric("RSI (9)", f"{latest_rsi:.2f}" if not pd.isna(latest_rsi) else "N/A")
        col3.metric("PRICE (3-EMA of RSI)", f"{float(latest['RSI_EMA_3']):.2f}" if not pd.isna(latest['RSI_EMA_3']) else "N/A")
        col4.metric("21-WMA of RSI", f"{float(latest['RSI_WMA_21']):.2f}" if not pd.isna(latest['RSI_WMA_21']) else "N/A")

        # Cap how much history feeds the chart: enough to scroll back a
        # good while, without letting the y-axis stretch to a decade's
        # price range (which shrinks the recent candles to nothing).
        MAX_CHART_BARS = 300
        chart_data = data.tail(MAX_CHART_BARS) if len(data) > MAX_CHART_BARS else data

        st.plotly_chart(
            build_chart(
                chart_data,
                num_bars=num_bars,
                interval=interval
            ),
            use_container_width=True,
            config={
                "scrollZoom": True,      # pinch (two-finger) to zoom in/out
                "displayModeBar": False  # cleaner mobile view
            }
        )

        with st.expander("Raw Data (latest rows)"):
            _fo_symbols = get_fo_symbols()
            _raw = (
                data[["Close", "RSI_9", "RSI_EMA_3", "RSI_WMA_21"]]
                .rename(columns={"Close": "Last Traded Price"})
                .tail(15).sort_index(ascending=False)
            )
            _raw = _raw.reset_index().rename(columns={_raw.index.name or "index": "Date"})
            # Show time too for intraday intervals; date-only for daily/weekly/monthly
            _is_intraday_raw = interval in ["1m", "5m", "10m", "15m", "30m", "1h", "2h", "3h", "4h"]
            _date_fmt = "%Y-%m-%d %H:%M" if _is_intraday_raw else "%Y-%m-%d"
            _raw["Date"] = pd.to_datetime(_raw["Date"]).dt.strftime(_date_fmt)
            _raw.insert(0, "Symbol", symbol)
            _raw["Category"] = "Index" if symbol.startswith("^") else get_category(symbol, _fo_symbols)

            render_pinned_table(_raw, ["Symbol", "Date"], pin_widths=[100, 160])

        st.caption(f"Last updated: {pd.Timestamp.now()}")

    except Exception as e:
        st.error(f"Error fetching data: {e}")


# ============================================================
# ---------- Market Scanner: Indexes + Nifty 500 ----------
# ============================================================

@st.cache_data(ttl=86400, show_spinner=False)
def get_fo_symbols():
    """Fetch the live list of F&O (Futures & Options) eligible symbols
    from NSE. Returns a set of bare symbols (no .NS). Empty set on failure
    (in which case every stock will show as 'Cash' until it succeeds)."""
    url = "https://nsearchives.nseindia.com/content/fo/fo_mktlots.csv"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0 Safari/537.36",
        "Accept-Language": "en-US,en;q=0.9",
    }
    try:
        session = requests.Session()
        session.headers.update(headers)
        session.get("https://www.nseindia.com", timeout=10)
        resp = session.get(url, timeout=15)
        resp.raise_for_status()
        df = pd.read_csv(io.StringIO(resp.text))
        df.columns = [c.strip() for c in df.columns]
        syms = df["SYMBOL"].astype(str).str.strip().tolist()
        return set(syms)
    except Exception:
        return set()


def get_category(symbol, fo_symbols):
    """F&O if the (bare, no .NS/^) symbol trades in Futures & Options, else Cash."""
    bare = symbol.replace(".NS", "").lstrip("^")
    return "F&O" if bare in fo_symbols else "Cash"


@st.cache_data(ttl=86400, show_spinner=False)
def get_nifty500_symbols():
    """Fetch the live Nifty 500 constituent list directly from NSE.
    Falls back to a small known list if NSE blocks the request."""
    symbols, _ = get_nifty500_details()
    return symbols


# Market-cap based category files, in the display/sort order requested
CATEGORY_FILES = [
    ("NIFTY 50", "ind_nifty50list.csv"),
    ("NIFTY NEXT 50", "ind_niftynext50list.csv"),
    ("NIFTY MIDCAP 150", "ind_niftymidcap150list.csv"),
    ("NIFTY SMALLCAP 250", "ind_niftysmallcap250list.csv"),
]
CATEGORY_ORDER = [c[0] for c in CATEGORY_FILES] + ["OTHER (NIFTY 500)"]


@st.cache_data(ttl=3600, show_spinner=False)
def get_stock_categories():
    """Fetch each market-cap category's live constituent list from NSE and
    build a symbol -> category lookup. If a stock appears in more than one
    list, the higher (larger market-cap) category wins."""
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0 Safari/537.36",
        "Accept-Language": "en-US,en;q=0.9",
        "Referer": "https://www.nseindia.com/",
    }
    category_map = {}
    category_symbol_order = {}
    fetch_errors = {}

    for name, fname in CATEGORY_FILES:
        try:
            # Fresh session per file -- NSE can reject a session's 2nd/3rd
            # request even though the cookie looks valid.
            session = requests.Session()
            session.headers.update(headers)
            session.get("https://www.nseindia.com", timeout=10)
            url = f"https://nsearchives.nseindia.com/content/indices/{fname}"
            resp = session.get(url, timeout=15)
            resp.raise_for_status()
            df = pd.read_csv(io.StringIO(resp.text))
            syms = df["Symbol"].astype(str).str.strip().tolist()
            if not syms:
                fetch_errors[name] = "File fetched but had no Symbol column / was empty"
        except Exception as e:
            syms = []
            fetch_errors[name] = str(e)
        category_symbol_order[name] = syms
        for s in syms:
            if s not in category_map:  # first (highest-priority) category wins
                category_map[s] = name

    return category_map, category_symbol_order, fetch_errors


def sort_stocks_by_category(df, category_map, category_symbol_order):
    """Sort rows: NIFTY 50 block first (in NSE's published list order),
    then NIFTY NEXT 50, then MIDCAP 250, then SMALLCAP 250, then leftovers."""
    def rank_key(symbol):
        cat = category_map.get(symbol, "OTHER (NIFTY 500)")
        cat_rank = CATEGORY_ORDER.index(cat) if cat in CATEGORY_ORDER else len(CATEGORY_ORDER)
        within_list = category_symbol_order.get(cat, [])
        sym_rank = within_list.index(symbol) if symbol in within_list else 9999
        return (cat_rank, sym_rank)

    df = df.copy()
    df["_sort_key"] = df["Symbol"].apply(rank_key)
    df = df.sort_values("_sort_key").drop(columns=["_sort_key"]).reset_index(drop=True)
    return df


# Shared signal -> color mapping (hex, used for both on-screen table and Excel)
SIGNAL_COLORS_EXCEL = {
    "BUY": "#008000",                        # green
    "SELL": "#FF0000",                       # red
    "CORRECTION/SHORT COVERING": "#0000FF",  # blue
    "WAIT": "#000000",                       # black (in the downloaded Excel)
}

SIGNAL_COLORS_DISPLAY = {
    "BUY": "#00FF00",                        # bright green (visible on dark bg)
    "SELL": "#FF4B4B",                        # red
    "CORRECTION/SHORT COVERING": "#4DA6FF",   # blue
    "WAIT": "#FFFFFF",                        # white (so it's visible on the dashboard)
}


def get_scanner_signal(rsi, price, wma):
    """
    PRICE here = EMA(3) of RSI (same 0-100 scale as RSI and WMA-of-RSI,
    consistent with the main chart's RSI-based signal lines).
    BUY: RSI > 50 and RSI > PRICE > WMA
    SELL: RSI < 50 and WMA > PRICE > RSI
    CORRECTION/SHORT COVERING: RSI < 50 and RSI > PRICE > WMA
    """
    if pd.isna(rsi) or pd.isna(price) or pd.isna(wma):
        return "WAIT"
    if rsi > 50 and rsi > price > wma:
        return "BUY"
    elif rsi < 50 and wma > price > rsi:
        return "SELL"
    elif rsi < 50 and rsi > price > wma:
        return "CORRECTION/SHORT COVERING"
    else:
        return "WAIT"


def render_pinned_table(df, pin_cols, row_color_col=None, pin_widths=None):
    """Render a dataframe as a plain HTML table with the given leading
    columns 'stuck' to the left using CSS position:sticky -- this works
    reliably regardless of the Streamlit version's dataframe-grid support
    (which turned out not to pin columns consistently). Cell text is also
    natively selectable/copyable this way.
    If row_color_col is given (e.g. "Signal"), each row's text color comes
    from SIGNAL_COLORS_DISPLAY based on that column's value.
    pin_widths: explicit pixel width per pinned column -- the width MUST
    match the actual rendered cell width exactly, or a gap opens up between
    pinned columns that scrolled content shows through."""
    import html as _html

    cols = list(df.columns)
    n_pin = len(pin_cols)
    # pin_cols must be the first columns, in order
    ordered = pin_cols + [c for c in cols if c not in pin_cols]
    df = df[ordered]
    numeric_cols = set(df.select_dtypes(include="number").columns)

    if pin_widths is None:
        pin_widths = [130] * n_pin

    def fmt(v, col):
        if pd.isna(v):
            return ""
        if col in numeric_cols:
            try:
                return f"{float(v):.2f}"
            except Exception:
                return str(v)
        return str(v)

    def pin_style(i):
        """Force the exact width so it matches the left-offset math below --
        a mismatch here is what let scrolled content bleed through."""
        w = pin_widths[i]
        return f"width:{w}px; min-width:{w}px; max-width:{w}px; overflow:hidden; text-overflow:ellipsis;"

    parts = [
        '<div style="overflow-x:auto; max-width:100%; border:1px solid #333; border-radius:4px;">',
        # border-collapse:collapse breaks position:sticky (collapsed borders
        # let scrolled content show through pinned cells) -- separate +
        # zero spacing avoids that while still looking like one solid grid.
        '<table style="border-collapse:separate; border-spacing:0; font-size:13px; color:#eee; table-layout:fixed;">',
        "<thead><tr>"
    ]
    left = 0
    for i, c in enumerate(ordered):
        style = ("padding:6px 10px; border:1px solid #333; background-color:#161b22; "
                 "white-space:nowrap; text-align:left;")
        if i < n_pin:
            style += f"position:sticky; left:{left}px; z-index:3; {pin_style(i)}"
            left += pin_widths[i]
        parts.append(f'<th style="{style}">{_html.escape(str(c))}</th>')
    parts.append("</tr></thead><tbody>")

    for _, row in df.iterrows():
        color = None
        if row_color_col:
            color = SIGNAL_COLORS_DISPLAY.get(row.get(row_color_col, "WAIT"), "#FFFFFF")
        color_style = f"color:{color};" if color else "color:#eee;"
        parts.append("<tr>")
        left = 0
        for i, c in enumerate(ordered):
            bg = "#1a1f27" if i < n_pin else "#0e1117"
            style = f"padding:6px 10px; border:1px solid #333; white-space:nowrap; background-color:{bg}; {color_style}"
            if i < n_pin:
                style += f"position:sticky; left:{left}px; z-index:2; {pin_style(i)}"
                left += pin_widths[i]
            parts.append(f'<td style="{style}">{_html.escape(fmt(row[c], c))}</td>')
        parts.append("</tr>")
    parts.append("</tbody></table></div>")
    st.markdown("".join(parts), unsafe_allow_html=True)


def style_signal_rows(df):
    """Colour every row's text based on its Signal column, and show all
    numeric values rounded to exactly 2 decimal places (for on-screen display)."""
    def _apply(row):
        color = SIGNAL_COLORS_DISPLAY.get(row.get("Signal", "WAIT"), "#FFFFFF")
        return [f"color: {color}"] * len(row)

    numeric_cols = df.select_dtypes(include="number").columns
    fmt = {col: "{:.2f}" for col in numeric_cols}
    return df.style.apply(_apply, axis=1).format(fmt)


def scan_symbols(symbols, yf_suffix=".NS", interval="1d"):
    """Batch-fetch data (at the chosen timeframe) and compute Latest Price,
    RSI(9), EMA(3) of RSI, WMA(21) of RSI and Signal for each symbol."""
    cfg = INTERVAL_CONFIG[interval]
    tickers = [
        f"{s}{yf_suffix}" if yf_suffix and not str(s).startswith("^") else s
        for s in symbols
    ]
    weights = np.arange(1, 22)

    try:
        raw = yf.download(
            tickers=" ".join(tickers),
            period=cfg["period"],
            interval=cfg["yf_interval"],
            group_by="ticker",
            threads=True,
            progress=False,
        )
    except Exception as e:
        st.error(f"Scanner fetch failed: {e}")
        return pd.DataFrame()

    rows = []
    for sym, tkr in zip(symbols, tickers):
        try:
            if len(tickers) == 1:
                df = raw
            else:
                if tkr not in raw.columns.get_level_values(0):
                    continue
                df = raw[tkr]

            df = df.dropna(how="all")

            # Combine smaller candles into the requested bigger duration
            # (same resample logic as the main single-stock chart, anchored
            # to 9:15 AM market open so buckets aren't uneven/partial)
            if cfg["resample"]:
                df = df.resample(cfg["resample"], offset="9h15min").agg({
                    "Open": "first", "High": "max", "Low": "min",
                    "Close": "last", "Volume": "sum"
                }).dropna(how="all")

            close = df["Close"].dropna()
            if len(close) < 10:
                continue

            delta = close.diff()
            gain = delta.clip(lower=0)
            loss = -delta.clip(upper=0)
            avg_gain = gain.ewm(alpha=1 / 9, min_periods=9, adjust=False).mean()
            avg_loss = loss.ewm(alpha=1 / 9, min_periods=9, adjust=False).mean()
            rs = avg_gain / avg_loss
            rsi = 100 - (100 / (1 + rs))

            rsi_ema3 = rsi.ewm(span=3, adjust=False).mean()
            rsi_wma21 = rsi.rolling(21).apply(
                lambda v: np.dot(v, weights) / weights.sum(), raw=True
            )

            latest_rsi = rsi.iloc[-1]
            latest_ema = rsi_ema3.iloc[-1]
            latest_wma = rsi_wma21.iloc[-1]
            signal = get_scanner_signal(latest_rsi, latest_ema, latest_wma)

            rows.append({
                "Symbol": sym,
                "Latest Price": round(float(close.iloc[-1]), 2),
                "RSI (9)": round(float(latest_rsi), 2) if not pd.isna(latest_rsi) else None,
                "EMA(3) of RSI": round(float(latest_ema), 2) if not pd.isna(latest_ema) else None,
                "WMA(21) of RSI": round(float(latest_wma), 2) if not pd.isna(latest_wma) else None,
                "Signal": signal,
            })
        except Exception:
            continue

    return pd.DataFrame(rows)


st.divider()
st.header("🔍 Market Scanner — Indexes & Nifty 500")
st.caption(
    "Separate from the live auto-refresh chart above. Scanning 500+ stocks "
    "can take a few minutes and only refreshes when you click the button "
    "(continuous live scanning of 500 stocks would get rate-limited by "
    "Yahoo Finance)."
)

scanner_interval = st.selectbox(
    "Scanner Timeframe",
    options=list(INTERVAL_CONFIG.keys()),
    index=list(INTERVAL_CONFIG.keys()).index("1d"),
    help="Data (and the Excel download) will be based on this timeframe."
)

if "scanner_index_df" not in st.session_state:
    st.session_state.scanner_index_df = None
if "scanner_stocks_df" not in st.session_state:
    st.session_state.scanner_stocks_df = None
if "scanner_interval_used" not in st.session_state:
    st.session_state.scanner_interval_used = None

if st.button("▶️ Run / Refresh Scanner"):
    with st.spinner("Fetching F&O eligible stock list from NSE..."):
        fo_symbols = get_fo_symbols()

    with st.spinner("Scanning major indexes..."):
        index_names = list(MAJOR_INDEXES.keys())
        index_symbols = list(MAJOR_INDEXES.values())
        idx_df = scan_symbols(index_symbols, yf_suffix="", interval=scanner_interval)
        if not idx_df.empty:
            name_map = dict(zip(index_symbols, index_names))
            # "Yahoo Name" holds the actual working ticker (copy this into the
            # sidebar symbol box). "Symbol" holds the friendly name so it's
            # easy to identify the row while it's pinned during scrolling.
            idx_df["Yahoo Name"] = idx_df["Symbol"]
            idx_df["Symbol"] = idx_df["Symbol"].map(name_map)
            # Indexes aren't individually "F&O" or "Cash" stocks -- they're
            # benchmarks with their own separate index-futures contracts.
            idx_df["Category"] = "Index"
            idx_df.insert(0, "Sr. No.", range(1, len(idx_df) + 1))
        st.session_state.scanner_index_df = idx_df

    with st.spinner("Fetching Nifty 50 / Next 50 / Midcap 150 / Smallcap 250 category lists..."):
        category_map, category_symbol_order, category_fetch_errors = get_stock_categories()
        for cat_name, err in category_fetch_errors.items():
            st.warning(f"⚠️ Could not fetch '{cat_name}' list from NSE ({err}). "
                       f"Those stocks will show as 'OTHER (NIFTY 500)' until this succeeds.")

    with st.spinner("Scanning Nifty 500 stocks (this can take a few minutes)..."):
        nifty500, nifty500_names = get_nifty500_details()
        stocks_df = scan_symbols(nifty500, yf_suffix=".NS", interval=scanner_interval)
        if not stocks_df.empty:
            stocks_df = sort_stocks_by_category(stocks_df, category_map, category_symbol_order)
            stocks_df.insert(0, "Market Cap Category", stocks_df["Symbol"].map(
                lambda s: category_map.get(s, "OTHER (NIFTY 500)")
            ))
            # F&O vs Cash, based on the bare symbol (before it's overwritten
            # with the company name below)
            fo_cash_col = stocks_df["Symbol"].map(lambda s: get_category(s, fo_symbols))
            stocks_df.insert(0, "Sr. No.", range(1, len(stocks_df) + 1))
            # "Yahoo Name" holds the actual working ticker (with .NS) to copy
            # into the sidebar symbol box. "Symbol" holds the company name.
            stocks_df["Yahoo Name"] = stocks_df["Symbol"] + ".NS"
            stocks_df["Symbol"] = stocks_df["Symbol"].map(nifty500_names).fillna(stocks_df["Symbol"])
            stocks_df["Category"] = fo_cash_col
            # Column order: Sr. No., Symbol (frozen pair), Market Cap Category,
            # data..., Yahoo Name, Category (F&O/Cash) -- last
            cols = ["Sr. No.", "Symbol", "Market Cap Category"] + [
                c for c in stocks_df.columns
                if c not in ("Sr. No.", "Symbol", "Market Cap Category", "Yahoo Name", "Category")
            ] + ["Yahoo Name", "Category"]
            stocks_df = stocks_df[cols]
        st.session_state.scanner_stocks_df = stocks_df

    st.session_state.scanner_interval_used = scanner_interval


def shorten_name(name, max_words=4):
    """Trim a long company/index name to at most `max_words` words for
    on-screen display (the full name still goes into the Excel export
    unchanged). Drops a trailing joining word like 'of'/'and' if it lands
    as the last word after trimming."""
    if not isinstance(name, str):
        return name
    stopwords = {"of", "and", "the", "for", "in", "&"}
    words = name.split()[:max_words]
    if words and words[-1].strip(".,").lower() in stopwords:
        words = words[:-1]
    return " ".join(words)


def display_frozen(df, name_col="Symbol"):
    """Show 'Sr. No.' and the Symbol column pinned while the rest of the
    columns (including Yahoo Name at the end) stay horizontally scrollable.
    Long names are shortened for display only -- the underlying data
    (used for the Excel download) is left untouched."""
    df2 = df.copy()
    df2[name_col] = df2[name_col].apply(shorten_name)
    render_pinned_table(df2, ["Sr. No.", name_col], row_color_col="Signal", pin_widths=[70, 170])


if st.session_state.scanner_index_df is not None and not st.session_state.scanner_index_df.empty:
    st.caption(f"Data timeframe: **{st.session_state.scanner_interval_used}**")

    tab_all, tab_idx, tab_stocks = st.tabs(["All", "Indexes", "Nifty 500 Stocks"])

    with tab_idx:
        display_frozen(st.session_state.scanner_index_df)

    with tab_stocks:
        display_frozen(st.session_state.scanner_stocks_df)

    with tab_all:
        st.write("**Indexes**")
        display_frozen(st.session_state.scanner_index_df)
        st.write("**Nifty 500 Stocks**")
        display_frozen(st.session_state.scanner_stocks_df)

    # ---- Download as a single Excel file with 2 separate sheets ----
    from openpyxl.styles import Font

    def color_code_sheet(worksheet, df, freeze_cols=2):
        if "Signal" not in df.columns:
            return
        numeric_col_idxs = [
            i + 1 for i, col in enumerate(df.columns)
            if pd.api.types.is_numeric_dtype(df[col])
        ]
        for row_idx, signal in enumerate(df["Signal"], start=2):  # row 1 = header
            color = SIGNAL_COLORS_EXCEL.get(signal, "#000000").lstrip("#")
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = Font(color=color)
                if col_idx in numeric_col_idxs:
                    cell.number_format = "0.00"
        # Freeze the first `freeze_cols` columns (Sr. No. + Symbol/Index) and header row
        from openpyxl.utils import get_column_letter
        freeze_cell = f"{get_column_letter(freeze_cols + 1)}2"
        worksheet.freeze_panes = freeze_cell

    excel_buffer = io.BytesIO()
    with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
        st.session_state.scanner_index_df.to_excel(writer, sheet_name="Indexes", index=False)
        st.session_state.scanner_stocks_df.to_excel(writer, sheet_name="Nifty500_Stocks", index=False)
        color_code_sheet(writer.sheets["Indexes"], st.session_state.scanner_index_df, freeze_cols=2)
        color_code_sheet(writer.sheets["Nifty500_Stocks"], st.session_state.scanner_stocks_df, freeze_cols=2)
    excel_buffer.seek(0)

    st.download_button(
        label="⬇️ Download Excel (Indexes + Nifty 500 — 2 sheets)",
        data=excel_buffer,
        file_name=(
            f"market_scan_{st.session_state.scanner_interval_used}_"
            f"{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.xlsx"
        ),
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
else:
    st.info("Click 'Run / Refresh Scanner' to load index and Nifty 500 data.")


# ---------- Live loop ----------
placeholder = st.empty()

if run:
    while True:
        with placeholder.container():
            fetch_and_display()
        time.sleep(refresh_seconds)
        st.rerun()
else:
    fetch_and_display()
